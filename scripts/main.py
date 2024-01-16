
from openpyxl.styles import PatternFill
from transformers import pipeline
from news import News
from summarize import Summarize
import pandas as pd
import os

MAX_LEN_FOR_PEGASUS = 200


def create_output_array(monitored_tickers, summaries, scores, urls):
    output_array = []
    for ticker in monitored_tickers:
        for i in range(len(summaries[ticker])):
            curr_output = [ticker,
                           summaries[ticker][i],
                           scores[ticker][i]['label'],
                           scores[ticker][i]['score'],
                           urls[ticker][i]]
            output_array.append(curr_output)
    return output_array


def write_to_terminal(monitored_tickers, final_output_array):
    headers = final_output_array[0]
    pd.set_option('display.max_columns', None)  # Display all columns
    pd.set_option('display.max_colwidth', 100)  # Display full column width
    for ticker in monitored_tickers:
        filtered_rows = [row for row in final_output_array if row[0] == ticker]
        df = pd.DataFrame(filtered_rows, columns=headers)

        print("\nAnalysis for {}".format(ticker))
        print(df)


def write_to_csv(monitored_tickers, final_output_array, filename):
    headers = final_output_array[0]
    os.makedirs("../analyzed_sentiment_files", exist_ok=True)

    # Construct the full path with the subdirectory
    full_path = os.path.join("../analyzed_sentiment_files", filename)
    with pd.ExcelWriter(filename) as writer:
        for ticker in monitored_tickers:
            filtered_rows = [row for row in final_output_array if row[0] == ticker]
            df = pd.DataFrame(filtered_rows, columns=headers)
            df.to_excel(writer, sheet_name=ticker, index=False)

            # Get the ExcelWriter workbook and sheet
            # workbook = writer.book
            worksheet = writer.sheets[ticker]

            # Get the column index for the 'Label' column
            label_column_idx = headers.index('Label')

            # Apply color based on sentiment
            for idx, row in df.iterrows():
                label = row['Label']
                cell = worksheet.cell(row=idx + 2, column=label_column_idx + 1)  # Adding 1 for 1-based indexing
                fill_color = '00FF00' if label == 'POSITIVE' else 'FF0000'  # Green for positive, Red for negative
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')



def main():
    user_input_tickers = input("Enter the tickers you want to monitor (comma-separated): ")
    monitored_tickers = [ticker.strip().upper() for ticker in user_input_tickers.split(',')]

    # Setting up models
    print("Setting up models")
    summarizer = Summarize()
    summarizer.set_up_model()
    news_model = News()

    print("Finished setting up models")
    raw_urls = {ticker: news_model.search_for_news(ticker) for ticker in monitored_tickers}
    # Cleaning up urls
    exclude_items = ['policies', 'maps', 'preferences', 'accounts', 'support']

    print("Cleaning up urls")
    cleaned_urls = {ticker: news_model.clean_urls(raw_urls[ticker], exclude_items, ticker)
                    for ticker in monitored_tickers}
    print("Done cleaning urls")
    # print(cleaned_urls)

    print("Processing and scraping news links")

    articles = {ticker: news_model.process(cleaned_urls[ticker], MAX_LEN_FOR_PEGASUS=MAX_LEN_FOR_PEGASUS, ticker=ticker) for ticker in
                monitored_tickers}

    print("Summarizing all the articles")

    summaries = {ticker: summarizer.summarize_all(articles[ticker], ticker) for ticker in monitored_tickers}

    print("Calculating sentiment for each summary")
    sentiment = pipeline('sentiment-analysis')
    scores = {ticker: sentiment(summaries[ticker]) for ticker in monitored_tickers}
    print("Finished calculating sentiment for each summary")
    print("Consolidating results")

    final_output_array = create_output_array(monitored_tickers, summaries, scores, cleaned_urls)
    final_output_array.insert(0, ['Ticker', 'Summary', 'Label', 'Confidence', 'URL'])

    print("Writing results to terminal")

    write_to_terminal(monitored_tickers, final_output_array)

    printed_to_file = input(
        "Do you want to print these results to a file? If so, write y, [name of file.xlsx] e.g. y, output.xlsx: \n")

    if printed_to_file[0] == 'y':
        if ',' not in printed_to_file:
            printed_to_file = input(
                "Please write in this format with the comma: y, output.xlsx: \n")
        write_to_csv(monitored_tickers, final_output_array, printed_to_file.split(',')[1])


if __name__ == "__main__":
    main()
